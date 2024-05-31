import openpyxl as excel
import openpyxl.styles as styles


class Style:
    def __init__(self, tip: str):
        self.tip = tip
        # za naslov
        if tip.lower() == "naslov":
            self.font = styles.Font(
                name='Arial', size=20, bold=True, color='000000')
            self.fill = styles.PatternFill(
                start_color='AA88FF', end_color='AA88FF', fill_type='solid')
            thick = styles.Side("thick")
            self.border = styles.Border(
                left=thick, right=thick, top=thick, bottom=thick)
            self.alignment = styles.Alignment(
                horizontal='center', vertical='center')
        # za vpodnaslov
        elif tip.lower() == "podnaslov":
            self.font = styles.Font(name='Arial', size=12, color='000000')
            self.fill = styles.PatternFill()
            thin = styles.Side("thin")
            self.border = styles.Border(left=thin, right=thin, bottom=thin)
            self.alignment = styles.Alignment(
                horizontal="center", vertical="center")
        # za vse ostalo
        else:
            self.font = styles.Font(name='Calibri', size=12, color='000000')
            self.fill = None
            self.border = None
            self.alignment = styles.Alignment(
                horizontal='left', vertical='center')


class CellCoord:
    def __init__(self, row: int, column: int) -> None:
        self.row = row
        self.col = column

    def __eq__(self, other) -> bool:
        return self.col == other.col and self.row == other.row

    def __ne__(self, other) -> bool:
        return self.col != other.col or self.row != other.row

    def format(self) -> str:
        return f"{excel.utils.get_column_letter(self.col)}{self.row}"

    def tuple(self) -> tuple[int, int]:
        return (self.row, self.col)


# list[row: int, column: int]
def CellRange(start_cell: CellCoord, end_cell: CellCoord) -> str:
    start_cell = excel.utils.get_column_letter(
        start_cell.col) + str(start_cell.row)
    end_cell = excel.utils.get_column_letter(end_cell.col) + str(end_cell.row)
    cell_range = f"{start_cell}:{end_cell}"
    return cell_range


# najde prvo (prosto) celico v stolpcu
def NajdiCelicoStolpec(tabela, zacetna_celica: CellCoord, vrednost=None) -> CellCoord:
    while tabela[zacetna_celica.format()].value != vrednost:
        zacetna_celica.row += 1
    return zacetna_celica


# najde prvo (prosto) celico v stolpcu
def NajdiCelicoVrsta(tabela, zacetna_celica: CellCoord, vrednost=None, absolutna_meja=1000) -> CellCoord:
    while tabela[zacetna_celica.format()].value != vrednost and zacetna_celica.col < absolutna_meja:
        print(f"{zacetna_celica.row}, {zacetna_celica.col}")
        zacetna_celica.col += 1
    if zacetna_celica.col == absolutna_meja:
        return CellCoord(1, 1)
    return zacetna_celica


def ApplyStyleToCells(sheet, start_row, start_col, end_row, end_col, style: Style) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = style.font
            cell.fill = style.fill
            cell.border = style.border
            cell.alignment = style.alignment
