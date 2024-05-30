import openpyxl as excel
import openpyxl.styles as styles
from UrediCelice import *


def NaslovPridevnik(tabela, cursor: CellCoord, naslov: Style, podnaslov: Style) -> None:
    tabela[cursor.format()] = "PRIDEVNIKI"
    tabela.merge_cells(CellRange(cursor, CellCoord(cursor.row, cursor.col+1)))
    # nastavi slog:naslov
    ApplyStyleToCells(tabela, cursor.row, cursor.col,
                      cursor.row, cursor.col+1, naslov)
    cursor.row += 1
    tabela[cursor.format()] = "pridevnik"
    cursor.col += 1
    tabela[cursor.format()] = "pomen"
    # nastavi slog:podnaslov
    ApplyStyleToCells(tabela, cursor.row, cursor.col-1,
                      cursor.row, cursor.col, podnaslov)
    # nastavi cursor na začetek naslednjega naslova (vrstica 3, 2 stolpca v desno)
    cursor.row -= 1
    cursor.col += 2


def NaslovGlagol(tabela, cursor: CellCoord, naslov: Style, podnaslov: Style) -> None:
    tabela[cursor.format()] = "GLAGOLI"
    tabela.merge_cells(CellRange(cursor, CellCoord(cursor.row, cursor.col+3)))
    # nastavi slog:naslov
    ApplyStyleToCells(tabela, cursor.row, cursor.col,
                      cursor.row, cursor.col+3, naslov)
    cursor.row += 1
    tabela[cursor.format()] = "glagol"
    cursor.col += 1
    tabela[cursor.format()] = "3. oseba"
    cursor.col += 1
    tabela[cursor.format()] = "perfekt?"
    cursor.col += 1
    tabela[cursor.format()] = "pomen"
    # nastavi slog:podnaslov
    ApplyStyleToCells(tabela, cursor.row, cursor.col-3,
                      cursor.row, cursor.col, podnaslov)
    # nastavi cursor na začetek naslednjega naslova (vrstica 3, 2 stolpca v desno)
    cursor.row -= 1
    cursor.col += 2


def NaslovPrislov(tabela, cursor: CellCoord, naslov: Style, podnaslov: Style) -> None:
    tabela[cursor.format()] = "PRISLOVI"
    tabela.merge_cells(CellRange(cursor, CellCoord(cursor.row, cursor.col+1)))
    # nastavi slog:naslov
    ApplyStyleToCells(tabela, cursor.row, cursor.col,
                      cursor.row, cursor.col+1, naslov)
    cursor.row += 1
    tabela[cursor.format()] = "prislov"
    cursor.col += 1
    tabela[cursor.format()] = "pomen"
    # nastavi slog:podnaslov
    ApplyStyleToCells(tabela, cursor.row, cursor.col-1,
                      cursor.row, cursor.col, podnaslov)
    # nastavi cursor na začetek naslednjega naslova (vrstica 3, 2 stolpca v desno)
    cursor.col += 2
    cursor.row -= 1


def NaslovDrugo(tabela, cursor: CellCoord, naslov: Style, podnaslov: Style) -> None:
    tabela[cursor.format()] = "DRUGE OBLIKE"
    tabela.merge_cells(CellRange(cursor, CellCoord(cursor.row, cursor.col+1)))
    # nastavi slog:naslov
    ApplyStyleToCells(tabela, cursor.row, cursor.col,
                      cursor.row, cursor.col+1, naslov)
    cursor.row += 1
    tabela[cursor.format()] = "besedna zveza"
    cursor.col += 1
    tabela[cursor.format()] = "pomen"
    # nastavi slog:podnaslov
    ApplyStyleToCells(tabela, cursor.row, cursor.col-1,
                      cursor.row, cursor.col, podnaslov)
    # nastavi cursor na začetek naslednjega naslova (vrstica 3, 2 stolpca v desno)
    cursor.row -= 1
    cursor.col += 2


def NovLetnik(workbook, trenuten_letnik: int) -> None:
    workbook.create_sheet(title=f"{trenuten_letnik+1}.letnik")
    nov_letnik = workbook[f"{trenuten_letnik+1}.letnik"]

    slog_naslov = Style("naslov")
    slog_podnaslov = Style("podnaslov")
    nov_letnik["A1"] = "Letnik:"
    nov_letnik["B1"] = trenuten_letnik+1
    cursor = CellCoord(3, 2)

    NaslovPridevnik(nov_letnik, cursor, slog_naslov, slog_podnaslov)
    NaslovGlagol(nov_letnik, cursor, slog_naslov, slog_podnaslov)
    NaslovPrislov(nov_letnik, cursor, slog_naslov, slog_podnaslov)
    NaslovDrugo(nov_letnik, cursor, slog_naslov, slog_podnaslov)
